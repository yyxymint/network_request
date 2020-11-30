import requests
from bs4 import BeautifulSoup
import operator
import time
import re
import os

from openpyxl import Workbook
from openpyxl import load_workbook

# options = webdriver.ChromeOptions()
# options.add_argument('headless')
# options.add_argument('window-size=1920x1080')
# options.add_argument("disable-gpu")

import pyautogui
import requests
import selenium
import time

from bs4 import BeautifulSoup

from datetime import datetime


from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import ElementNotVisibleException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import UnexpectedAlertPresentException

def request(url):
    header = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9',
        'Accept-Encoding': 'gzip,deflate',
        'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Host': 'cancer.go.kr',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0;Win64;x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36'
    }
    try:
        url_get = requests.get(url, headers=header)
    except:
        url_get = requests.get(url, headers=header)
    return url_get
   
def get_soup(url):
	recept=request(url)
	return BeautifulSoup(recept.text, "html.parser")

def there_is_under(word):    #아스키(ASCII) 코드 공식에 따라 입력된 단어의 마지막 글자 받침 유무를 판단해서 뒤에 붙는 조사를 리턴하는 함수
    last = word[-1]     #입력된 word의 마지막 글자를 선택해서
    criteria = (ord(last) - 44032) % 28     #아스키(ASCII) 코드 공식에 따라 계산 (계산법은 다음 포스팅을 참고하였습니다 : http://gpgstudy.com/forum/viewtopic.php?p=45059#p45059)
    if criteria == 0:       #나머지가 0이면 받침이 없는 것
        return False
    else:                   #나머지가 0이 아니면 받침 있는 것
        return True

      
def close_tab():
	driver.switch_to.window(driver.window_handles[-1])
	driver.close()
	driver.switch_to.window(driver.window_handles[-1])

def back():
	driver.switch_to.window(driver.window_handles[-1])

def by_xp(s):
	driver.find_element_by_xpath(s).click()
  
# driver.find_element_by_xpath('//*[@id="id"]').send_keys(m_id)

# driver = webdriver.Chrome('C:/Users/HWNAG YE CHAN/Downloads/zoo/chromedriver.exe')
# driver.get('https://www.melon.com/index.htm')



# load_wb = load_workbook("C:\\Users\\HWNAG YE CHAN\\Desktop\\cancer74.xlsx", data_only=True)
# load_ws = load_wb['Sheet']

# # print(load_ws['A1'].value)

# write_wb = Workbook()
# write_ws = write_wb.active

# last_name="가성점액종"
# for row in range(1,1323,1):
# 	write_ws['A'+str(row)]=load_ws['A'+str(row)].value
# 	write_ws['B'+str(row)]=load_ws['B'+str(row)].value
# 	write_ws['C'+str(row)]="암정보"
# 	now_q=load_ws['A'+str(row)].value
# 	for q in range(len(longqlist)):
# 		s=now_q.find(longqlist[q])
# 		if s!=-1:
# 			now_cancer_name=now_q[:s]
# 			if now_cancer_name!=last_name:
# 				print("done "+last_name)
# 			last_name=now_cancer_name
# 			write_ws['E'+str(row)]=now_cancer_name
# 			write_ws['F'+str(row)]=appendlist[q]
# write_wb.save('edited.xlsx')



docker run \
    --detach \
    --gpus=all \
    --restart=always \
    --name=$(whoami)-jupyterlab \
    --publish=80 \
    --publish=6006 \
    --volume=/home/$(whoami)/notebooks:/notebooks \
    --volume=/storage/$(whoami):/storage \
    jonghwanhyeon/jupyterlab
