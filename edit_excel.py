import requests
from bs4 import BeautifulSoup
import operator
import time
import re
import os

from openpyxl import Workbook
from openpyxl import load_workbook

canlist=['가성점액종', '간내 담도암', '간모세포종', '간암', '갑상선암', '결장암', '고환암', '골수이형성증후군', '교모세포종', '구강암', '구순암', '구인두-편도암(HPV 관련)', '구인두-하인두암(HPV 비관련)', '균상식육종', '급성골수성백혈병', '급성림프구성백혈병', '기저세포암', '난소상피암', '난소생식세포종양', '남성유방암', '뇌종양', '뇌하수체선종', '다발성골수종', '담낭·담도암', '담낭암', '담도암', '대장암', '만성골수성백혈병', '만성림프구백혈병', '망막모세포종', '맥락막흑색종', '미만성거대B세포림프종', '바터팽대부암', '방광암', '복막암', '부갑상선암', '부신암', '비부비동암', '비소세포폐암', '비호지킨림프종', '설암', '성상세포종', '소세포폐암', '소아뇌종양', '소아림프종', '소아백혈병', '소장암', '수막종', '식도암', '신경교종', '신경모세포종', '신우암', '신장암', '심장암', '십이지장암', '악성 연부조직 종양', '악성골종양', '악성림프종', '악성중피종', '악성흑색종', '안종양', '외음부암', '요관암', '요도암', '원발부위불명암', '위림프종', '위암', '위유암종', '위장관기질종양', '윌름스종양', '유방암', '육종', '음경암', '임신융모질환', '자궁경부암', '자궁내막암', '자궁육종', '전립선암', '전이성 골종양', '전이성뇌종양', '종격동암', '직장암', '직장유암종', '질암', '척수종양', '청신경초종', '췌장암', '침샘암', '카포시 육종', '파제트병', '편평상피세포암', '폐선암', '폐암', '폐편평상피세포암', '피부암', '항문암', '횡문근육종', '후두암', '흉막암', '흉선암']
longqlist=[
'의 발생부위는 어디인가요?',
'은 무엇인가요?',
' 관련 통계가 있나요?',
'의 위험요인은 무엇인가요?',
'의 예방법은 무엇인가요?',
'의 조기검진 방법은 무엇인가요?',
'의 증상은 무엇인가요?',
'의 진단방법은 무엇인가요?',
'의 진행단계는 어떻게 되나요?',
'의 감별진단은 어떻게 하나요?',
'의 치료방법은 무엇인가요?',
' 치료의 부작용은 무엇인가요?',
'는 재발과 전이가 잘 일어나나요?',
'의 치료현황은 어떤가요?',
]

appendlist=[
'발생부위',
'정의',
'',
'암발병',
'암예방',
'',
'증상',
'진단',
'',
'감별진단',
'치료방법',
'부작용',
'전이',
''
]

load_wb = load_workbook("C:\\Users\\HWNAG YE CHAN\\Desktop\\cancer74.xlsx", data_only=True)
load_ws = load_wb['Sheet']

# print(load_ws['A1'].value)

write_wb = Workbook()
write_ws = write_wb.active

last_name="가성점액종"
for row in range(1,1323,1):
	write_ws['A'+str(row)]=load_ws['A'+str(row)].value
	write_ws['B'+str(row)]=load_ws['B'+str(row)].value
	write_ws['C'+str(row)]="암정보"
	now_q=load_ws['A'+str(row)].value
	for q in range(len(longqlist)):
		s=now_q.find(longqlist[q])
		if s!=-1:
			now_cancer_name=now_q[:s]
			if now_cancer_name!=last_name:
				print("done "+last_name)
			last_name=now_cancer_name
			write_ws['E'+str(row)]=now_cancer_name
			write_ws['F'+str(row)]=appendlist[q]
write_wb.save('edited.xlsx')
